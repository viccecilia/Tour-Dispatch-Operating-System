from __future__ import annotations

import re
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from backend.db.database import get_connection
from backend.services.order_number_service import normalize_vehicle_type_label
from backend.services.tenant_context import get_current_tenant_id


DEFAULT_LOCATIONS = [
    {"std_name": "KIX", "loc_type": "Airport", "aliases": "KIX,关西,关空,关西机场,关西空港,关西国际机场,Kansai Airport"},
    {"std_name": "ITM", "loc_type": "Airport", "aliases": "ITM,伊丹,伊丹机场,大阪伊丹机场"},
    {"std_name": "UKB", "loc_type": "Airport", "aliases": "UKB,神户机场,神户空港"},
    {"std_name": "HND", "loc_type": "Airport", "aliases": "HND,羽田,羽田机场,东京羽田"},
    {"std_name": "NRT", "loc_type": "Airport", "aliases": "NRT,成田,成田机场,东京成田"},
    {"std_name": "NGO", "loc_type": "Airport", "aliases": "NGO,名古屋机场,中部机场,中部国际机场"},
    {"std_name": "大阪市内", "loc_type": "City", "aliases": "大阪,大阪市内,Osaka,Osaka City,大阪酒店"},
    {"std_name": "京都市内", "loc_type": "City", "aliases": "京都,京都市内,Kyoto,Kyoto City,京都酒店"},
    {"std_name": "奈良", "loc_type": "City", "aliases": "奈良,Nara"},
    {"std_name": "宇治", "loc_type": "City", "aliases": "宇治,Uji"},
    {"std_name": "神户", "loc_type": "City", "aliases": "神户,Kobe"},
    {"std_name": "名古屋", "loc_type": "City", "aliases": "名古屋,Nagoya"},
    {"std_name": "铃鹿", "loc_type": "City", "aliases": "铃鹿,Suzuka"},
    {"std_name": "大津", "loc_type": "City", "aliases": "大津,Otsu"},
    {"std_name": "新大阪", "loc_type": "Station", "aliases": "新大阪,新大阪站"},
    {"std_name": "环球影城", "loc_type": "Spot", "aliases": "环球,环球影城,USJ,Universal Studios Japan"},
    {"std_name": "天桥立美山", "loc_type": "Spot", "aliases": "天桥立美山,天橋立美山,天桥立,天橋立,美山"},
    {"std_name": "龟冈", "loc_type": "Spot", "aliases": "龟冈,龜岡,Kameoka"},
    {"std_name": "胜尾寺", "loc_type": "Spot", "aliases": "胜尾寺,勝尾寺,Katsuoji"},
]

NOTE_TOKEN_LIBRARY = [
    ("儿童座椅", ["儿童座椅", "儿童椅", "安全座椅", "child seat", "baby seat"]),
    ("婴儿座椅", ["婴儿座椅", "婴儿椅", "baby seat"]),
    ("接机牌", ["接机牌", "举牌", "meet sign"]),
    ("深夜", ["深夜", "夜间", "早班"]),
    ("航班", ["航班", "航班号", "flight", "CA", "JL", "NH", "MU", "CZ"]),
    ("轮椅", ["轮椅", "wheelchair"]),
    ("代收", ["代收"]),
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}:
        return ""
    return (
        text.replace("\u3000", " ")
        .replace("，", ",")
        .replace("。", ".")
        .replace("－", "-")
        .replace("—", "-")
        .replace("→", "->")
        .strip()
    )


def upsert_locations(conn: sqlite3.Connection, locations: list[dict[str, Any]]) -> None:
    for loc in locations:
        std_name = clean_text(loc.get("std_name"))
        if not std_name:
            continue
        conn.execute(
            """
            INSERT INTO locations (std_name, loc_type, aliases, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(std_name) DO UPDATE SET
                loc_type = excluded.loc_type,
                aliases = excluded.aliases,
                updated_at = CURRENT_TIMESTAMP
            """,
            (std_name, clean_text(loc.get("loc_type")), clean_text(loc.get("aliases"))),
        )


def seed_default_locations(conn: sqlite3.Connection) -> None:
    upsert_locations(conn, DEFAULT_LOCATIONS)
    framework = Path(r"C:\PycharmProjects\pythonProject01\tourism interactive platform\Framework.xlsx")
    if framework.exists():
        upsert_locations(conn, load_locations_from_excel(framework))


def load_locations_from_excel(path: Path, sheet_name: str = "location-data") -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_text(cell) for cell in rows[0]]
    result = []
    for row in rows[1:]:
        data = {header[index]: row[index] if index < len(row) else None for index in range(len(header))}
        std_name = clean_text(data.get("标准名") or data.get("std_name") or data.get("名称") or data.get("地点"))
        if not std_name:
            continue
        result.append(
            {
                "std_name": std_name,
                "loc_type": clean_text(data.get("类型") or data.get("loc_type")),
                "aliases": clean_text(data.get("常用名") or data.get("aliases") or data.get("别名")),
            }
        )
    return result


def list_locations(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute("SELECT std_name, loc_type, aliases FROM locations ORDER BY std_name").fetchall()
    return [dict(row) for row in rows]


def load_alias_index(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    index = []
    for row in conn.execute("SELECT std_name, loc_type, aliases FROM locations").fetchall():
        aliases = [clean_text(item) for item in clean_text(row["aliases"]).split(",") if clean_text(item)]
        if row["std_name"] not in aliases:
            aliases.append(row["std_name"])
        index.append({"std_name": row["std_name"], "loc_type": row["loc_type"], "aliases": aliases})
    return index


def normalize_location_text(conn: sqlite3.Connection, raw_text: str | None) -> str:
    raw = clean_text(raw_text)
    if not raw:
        return ""
    compact = _compact(raw)
    preferred = {
        "关西": "KIX",
        "关空": "KIX",
        "关西机场": "KIX",
        "关西空港": "KIX",
        "大阪": "大阪市内",
        "大阪市内": "大阪市内",
        "osaka": "大阪市内",
        "京都": "京都市内",
        "京都市内": "京都市内",
        "环球": "环球影城",
        "usj": "环球影城",
    }
    for key, value in preferred.items():
        if compact == _compact(key):
            return value
    index = load_alias_index(conn)
    for rec in index:
        for alias in rec["aliases"]:
            if _compact(alias) == compact:
                return rec["std_name"]
    for rec in sorted(index, key=lambda item: max([len(a) for a in item["aliases"]] or [0]), reverse=True):
        for alias in sorted(rec["aliases"], key=len, reverse=True):
            alias_compact = _compact(alias)
            if alias_compact and alias_compact in compact:
                return rec["std_name"]
    return raw


def normalize_date_token(raw_text: str, base_year: int | None = None) -> str:
    base_year = base_year or date.today().year
    text = clean_text(raw_text)
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    text = text.replace("/", "-").replace(".", "-").replace("年", "-").replace("月", "-").replace("日", "")
    text = re.sub(r"-+", "-", text).strip("-")
    if re.fullmatch(r"\d{8}", text):
        return f"{int(text[:4]):04d}-{int(text[4:6]):02d}-{int(text[6:8]):02d}"
    if re.fullmatch(r"\d{6}", text):
        return f"20{int(text[:2]):02d}-{int(text[2:4]):02d}-{int(text[4:6]):02d}"
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"
    match = re.fullmatch(r"(\d{2})-(\d{1,2})-(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"20{year:02d}-{month:02d}-{day:02d}"
    match = re.fullmatch(r"(\d{1,2})-(\d{1,2})", text)
    if match:
        month, day = map(int, match.groups())
        return f"{base_year:04d}-{month:02d}-{day:02d}"
    return ""


def normalize_time_token(raw_text: str) -> str:
    text = clean_text(raw_text)
    if not text:
        return ""
    text = re.sub(r"\s+", "", text).replace(".", ":").replace("-", ":")
    if re.fullmatch(r"\d{4}", text):
        return f"{int(text[:2]):02d}:{int(text[2:4]):02d}"
    match = re.fullmatch(r"(\d{1,2}):(\d{1,2})", text)
    if match:
        hour, minute = map(int, match.groups())
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
    if re.fullmatch(r"\d{1,2}", text):
        hour = int(text)
        if 0 <= hour <= 23:
            return f"{hour:02d}:00"
    return ""


def identify_vehicle_type(text: str) -> str | None:
    raw = clean_text(text).lower()
    parts = []
    if any(token in raw for token in ["3代", "alphard", "埃尔法", "阿尔法", "vellfire"]):
        parts.append("3代")
    if any(token in raw for token in ["10座", "十座", "海狮", "hiace"]):
        parts.append("10座")
    if "18座" in raw or "中巴" in raw:
        parts.append("18座")
    if "绿牌" in raw or "绿" in raw:
        parts.append("绿牌")
    return " ".join(dict.fromkeys(parts)) or None


def identify_vehicle_type(text: str) -> str | None:
    raw = clean_text(text).lower()
    parts = []
    if any(token in raw for token in ["3代", "三代", "alphard", "埃尔法", "阿尔法", "アルファード", "vellfire", "ヴェルファイア"]):
        parts.append("3代")
    if any(token in raw for token in ["10座", "十座", "海狮", "海獅", "hiace", "ハイエース"]):
        parts.append("10座")
    if any(token in raw for token in ["7座", "七座", "gl8", "别克gl8"]):
        parts.append("7座")
    if any(token in raw for token in ["18座", "中巴", "マイクロバス"]):
        parts.append("18座")
    if any(token in raw for token in ["23座", "考斯特", "coaster"]):
        parts.append("23座")
    if any(token in raw for token in ["30系", "30アルファード", "30 alphard"]):
        parts.append("30系")
    if any(token in raw for token in ["绿牌", "绿", "绿色"]):
        parts.append("绿牌")
    if any(token in raw for token in ["白牌", "白色"]):
        parts.append("白牌")
    return " ".join(dict.fromkeys(parts)) or None


def identify_vehicle_type(text: str) -> str | None:
    label = normalize_vehicle_type_label(text)
    return label or None


def extract_note_tokens(raw_text: str) -> tuple[list[str], str]:
    raw = clean_text(raw_text)
    notes = []
    for label, aliases in NOTE_TOKEN_LIBRARY:
        if any(alias.lower() in raw.lower() for alias in aliases):
            notes.append(label)
    for match in re.finditer(r"代收\s*([0-9,]+)\s*(?:日元|jpy)?", raw, re.IGNORECASE):
        notes.append(f"代收{match.group(1)}日元")
    for match in re.finditer(r"(?:儿童座椅|儿童椅|安全座椅)\s*[*xX]?\s*(\d+)", raw, re.IGNORECASE):
        notes.append(f"儿童座椅*{match.group(1)}")
    for value in re.findall(r"（([^）]+)）|\(([^)]+)\)", raw):
        note = clean_text(value[0] or value[1])
        if note:
            notes.append(note)
    return list(dict.fromkeys(notes)), raw


def get_latest_locations(
    driver_id: Any = None,
    limit: int = 50,
    online_status: str | None = None,
    vehicle_status: str | None = None,
) -> list[dict[str, Any]]:
    params: list[Any] = [get_current_tenant_id()]
    where = ["ll.tenant_id = ?"]
    if driver_id not in ("", None):
        where.append("ll.driver_id = ?")
        params.append(_to_int(driver_id))
    params.append(max(1, min(int(limit or 50), 200)))
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM (
                SELECT
                    ll.*,
                    d.name AS driver_name,
                    d.phone AS driver_phone,
                    v.plate_number,
                    v.vehicle_type,
                    v.status AS vehicle_status,
                    o.oid,
                    o.pickup_location,
                    o.dropoff_location,
                    o.order_date,
                    o.start_time,
                    o.end_time,
                    o.dispatch_status,
                    o.settlement_status,
                    o.execution_status AS order_execution_status,
                    a.status AS assignment_status,
                    a.execution_status,
                    ROW_NUMBER() OVER (PARTITION BY ll.driver_id ORDER BY ll.reported_at DESC, ll.id DESC) AS rn
                FROM location_logs ll
                LEFT JOIN drivers d ON d.id = ll.driver_id
                LEFT JOIN vehicles v ON v.id = ll.vehicle_id
                LEFT JOIN assignments a ON a.id = ll.assignment_id
                LEFT JOIN orders o ON o.id = ll.order_id
                WHERE {" AND ".join(where)}
            )
            WHERE rn = 1
            ORDER BY reported_at DESC, id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()
    locations = [_with_online_status(dict(row)) for row in rows]
    if online_status:
        locations = [item for item in locations if item.get("online_status") == online_status]
    if vehicle_status:
        locations = [item for item in locations if item.get("vehicle_status") == vehicle_status]
    return locations


def list_location_logs(driver_id: Any = None, limit: int = 100) -> list[dict[str, Any]]:
    params: list[Any] = [get_current_tenant_id()]
    where = ["ll.tenant_id = ?"]
    if driver_id not in ("", None):
        where.append("ll.driver_id = ?")
        params.append(_to_int(driver_id))
    params.append(max(1, min(int(limit or 100), 300)))
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT
                ll.*,
                d.name AS driver_name,
                d.phone AS driver_phone,
                v.plate_number,
                v.vehicle_type,
                v.status AS vehicle_status,
                o.oid,
                o.pickup_location,
                o.dropoff_location,
                o.order_date,
                o.start_time,
                o.end_time,
                o.dispatch_status,
                o.settlement_status,
                o.execution_status AS order_execution_status,
                a.status AS assignment_status,
                a.execution_status
            FROM location_logs ll
            LEFT JOIN drivers d ON d.id = ll.driver_id
            LEFT JOIN vehicles v ON v.id = ll.vehicle_id
            LEFT JOIN assignments a ON a.id = ll.assignment_id
            LEFT JOIN orders o ON o.id = ll.order_id
            WHERE {" AND ".join(where)}
            ORDER BY ll.reported_at DESC, ll.id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()
    return [_with_online_status(dict(row)) for row in rows]


def get_fleet_location_summary() -> dict[str, Any]:
    locations = get_latest_locations(limit=200)
    return {
        "total": len(locations),
        "online": len([item for item in locations if item.get("online_status") == "online"]),
        "stale": len([item for item in locations if item.get("online_status") == "stale"]),
        "unknown": len([item for item in locations if item.get("online_status") == "unknown"]),
        "locations": locations,
    }


def _with_online_status(row: dict[str, Any]) -> dict[str, Any]:
    try:
        reported_at = datetime.fromisoformat(str(row.get("reported_at")).replace("Z", ""))
        row["online_status"] = "online" if datetime.utcnow() - reported_at <= timedelta(minutes=15) else "stale"
    except (TypeError, ValueError):
        row["online_status"] = "unknown"
    return row


def _to_int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _compact(value: str) -> str:
    return re.sub(r"[\s,，。/\\_\->]+", "", clean_text(value).lower())
