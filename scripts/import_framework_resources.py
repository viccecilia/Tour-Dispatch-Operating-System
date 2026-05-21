from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.db.database import get_connection, init_db
from backend.services.resource_service import _add_months, _add_years, _ensure_resource_columns

DEFAULT_FRAMEWORK = Path(r"C:\PycharmProjects\pythonProject01\tourism interactive platform\Framework.xlsx")
GREEN_RGB = {"FF92D050", "0092D050"}
DEFAULT_LIMIT = 15


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FRAMEWORK
    if not path.exists():
        raise SystemExit(f"Framework.xlsx not found: {path}")
    init_db(seed=False)
    _ensure_resource_columns()
    workbook = openpyxl.load_workbook(path, data_only=True)
    limit = None if "--all" in sys.argv else DEFAULT_LIMIT
    driver_count = import_drivers(workbook["drivers-data"], limit=limit)
    vehicle_count, record_count = import_vehicles(workbook["car-data"], limit=limit)
    print(
        {
            "framework": str(path),
            "drivers_upserted": driver_count,
            "vehicles_upserted": vehicle_count,
            "limit": limit or "all",
            "vehicle_inspection_records": record_count,
        }
    )


def import_drivers(ws, limit: int | None = DEFAULT_LIMIT) -> int:
    count = 0
    with get_connection() as conn:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if limit is not None and count >= limit:
                break
            if not row or not row[3]:
                continue
            payload = {
                "driver_external_id": clean(row[1]),
                "office": clean(row[2]),
                "name": clean(row[3]),
                "driver_code": clean(row[4]),
                "driver_language": clean(row[5]),
                "license_due_date": parse_date(row[6]),
                "license_expires_at": parse_date(row[6]),
                "license_number": clean(row[7]),
                "residence_status": clean(row[8]),
                "residence_due_date": parse_date(row[9]),
                "health_check_due_date": parse_date(row[10]),
                "medical_check_expires_at": parse_date(row[10]),
                "health_check_remaining_days": to_int(row[11]),
                "phone": clean(row[12]),
                "wechat": clean(row[13]),
                "line": clean(row[14]),
                "whatsapp": clean(row[15]),
                "email": clean(row[16]),
                "status": driver_status(row[17]),
                "driver_status": driver_status(row[17]),
            }
            existing = conn.execute(
                "SELECT id FROM drivers WHERE tenant_id = 1 AND (name = ? OR driver_external_id = ?)",
                (payload["name"], payload["driver_external_id"]),
            ).fetchone()
            if existing:
                assignments = ", ".join(f"{key} = ?" for key in payload)
                conn.execute(
                    f"UPDATE drivers SET {assignments}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    [*payload.values(), existing["id"]],
                )
            else:
                columns = ", ".join(["tenant_id", *payload.keys(), "updated_at"])
                placeholders = ", ".join(["?", *("?" for _ in payload), "CURRENT_TIMESTAMP"])
                conn.execute(
                    f"INSERT INTO drivers ({columns}) VALUES ({placeholders})",
                    [1, *payload.values()],
                )
            count += 1
        conn.commit()
    return count


def import_vehicles(ws, limit: int | None = DEFAULT_LIMIT) -> tuple[int, int]:
    vehicle_count = 0
    record_count = 0
    with get_connection() as conn:
        for row_index in range(6, ws.max_row + 1):
            if limit is not None and vehicle_count >= limit:
                break
            plate_number = clean(ws.cell(row_index, 3).value)
            if not plate_number:
                continue
            records = []
            for col_index in range(11, ws.max_column + 1):
                cell = ws.cell(row_index, col_index)
                parsed = parse_date(cell.value)
                if not parsed:
                    continue
                fill_rgb = str(cell.fill.fgColor.rgb)
                records.append(
                    {
                        "inspection_date": parsed,
                        "inspection_type": "shaken" if fill_rgb in GREEN_RGB else "inspection",
                        "source": "Framework.xlsx",
                        "note": clean(cell.value),
                    }
                )
            latest_any = latest_date(record["inspection_date"] for record in records)
            latest_shaken = latest_date(record["inspection_date"] for record in records if record["inspection_type"] == "shaken")
            payload = {
                "plate_no": plate_number,
                "plate_number": plate_number,
                "vehicle_type": clean(ws.cell(row_index, 2).value),
                "plate_short_code": clean(ws.cell(row_index, 4).value),
                "vehicle_type_code": clean(ws.cell(row_index, 5).value),
                "vehicle_color": clean(ws.cell(row_index, 6).value),
                "snow_tire": "yes" if clean(ws.cell(row_index, 7).value) else "no",
                "vehicle_group": clean(ws.cell(row_index, 8).value),
                "first_registration_date": parse_date(ws.cell(row_index, 9).value),
                "company_registration_date": parse_date(ws.cell(row_index, 10).value),
                "last_inspection_date": latest_any.isoformat() if latest_any else None,
                "next_inspection_due_date": _add_months(latest_any, 3).isoformat() if latest_any else None,
                "shaken_due_date": _add_years(latest_shaken, 1).isoformat() if latest_shaken else None,
                "inspection_expires_at": _add_months(latest_any, 3).isoformat() if latest_any else None,
                "seat_count": infer_seats(clean(ws.cell(row_index, 2).value)),
                "seats": infer_seats(clean(ws.cell(row_index, 2).value)),
                "status": "available",
            }
            existing = conn.execute("SELECT id FROM vehicles WHERE tenant_id = 1 AND plate_number = ?", (plate_number,)).fetchone()
            if existing:
                vehicle_id = existing["id"]
                assignments = ", ".join(f"{key} = ?" for key in payload)
                conn.execute(
                    f"UPDATE vehicles SET {assignments}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    [*payload.values(), vehicle_id],
                )
            else:
                columns = ", ".join(["tenant_id", *payload.keys(), "updated_at"])
                placeholders = ", ".join(["?", *("?" for _ in payload), "CURRENT_TIMESTAMP"])
                cursor = conn.execute(
                    f"INSERT INTO vehicles ({columns}) VALUES ({placeholders})",
                    [1, *payload.values()],
                )
                vehicle_id = cursor.lastrowid
            conn.execute("DELETE FROM vehicle_inspection_records WHERE tenant_id = 1 AND vehicle_id = ?", (vehicle_id,))
            for record in records:
                conn.execute(
                    """
                    INSERT INTO vehicle_inspection_records (
                        tenant_id, vehicle_id, inspection_type, inspection_date, source, note, updated_at
                    )
                    VALUES (1, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (vehicle_id, record["inspection_type"], record["inspection_date"], record["source"], record["note"]),
                )
            vehicle_count += 1
            record_count += len(records)
        conn.commit()
    return vehicle_count, record_count


def clean(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def to_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit() and len(text) == 8:
        return datetime.strptime(text, "%Y%m%d").date().isoformat()
    text = text.replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed.isoformat()
        except ValueError:
            pass
    era_match = __import__("re").match(r"([RrHh])\s*(\d{1,2})\s*-?\s*(\d{1,2})(\d{2})?$", text)
    if era_match:
        era, year_text, month_text, day_text = era_match.groups()
        base = 2018 if era.upper() == "R" else 1988
        year = base + int(year_text)
        month = int(month_text)
        day = int(day_text or "1")
        return date(year, month, day).isoformat()
    return None


def latest_date(values) -> date | None:
    parsed = []
    for value in values:
        try:
            parsed.append(datetime.strptime(value, "%Y-%m-%d").date())
        except (TypeError, ValueError):
            pass
    return max(parsed) if parsed else None


def driver_status(value: Any) -> str:
    text = clean(value) or ""
    if "不可" in text or "停止" in text or "停用" in text:
        return "inactive"
    if "休" in text:
        return "resting"
    return "available"


def infer_seats(vehicle_type: str | None) -> int:
    text = vehicle_type or ""
    if "ハイエース" in text or "10" in text:
        return 10
    return 7


if __name__ == "__main__":
    main()
